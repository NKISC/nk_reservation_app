import datetime
import random
import sqlite3

import numpy as np

from backend import utils
from typing import *
from openpyxl import Workbook
from openpyxl.styles import *
import traceback
import matplotlib.pyplot as plt


def query_classroom(cond: dict[str, Any]) -> list[dict[str, Any]]:
    """
    Query the classroom information.
    :param cond: A dictionary containing the filters that select the classrooms to return.
        Possible keys:
            id (int): The classroom id.
            display (str): The display name of the classroom.
            place (str): The place of the classroom.
            pic_url (list[str]): The pic_url(s) of the classroom.
            func_tag (list[str]): The function tag(s) of the classroom.
        :return: A list of classroom information.
    """
    with sqlite3.connect('database.db') as db:
        cursor = db.cursor()
        param = utils.construct_params(cond)
        if len(cond) != 0:
            cursor.execute("select * from classroom where " + utils.construct_condition(cond), param)
        else:
            cursor.execute("select * from classroom")
        ret = utils.construct_response(cursor, "classroom")
        return ret


def query_record(cond: dict[str, Any]) -> list[dict[str, Any]]:
    """
    Query the reservation records.
    :param cond: A dictionary containing the filters that select the reservation records to return.
        Possible keys:
            id (int): The record id.
            classroom_id (int): The id of the classroom that is reserved in this record.
            noon (boolean): Whether the reservation is at noon.
            applicant_id (str): The applicant id.
            time_stamp (int): The date of the reservation (h, m, s, f are set to zero.
                              For instance, if the reservation is on Feb. 1, 2025, the time_stamp will be 1738339200).
    :return: A list of reservation information.
    """
    with sqlite3.connect('database.db') as db:
        utils.update_record()
        cursor = db.cursor()
        c = utils.construct_condition(cond)
        if len(cond) != 0 and c != "":
            cursor.execute("select * from record where " + c, cond)
        else:
            cursor.execute("select * from record")
        ret = utils.construct_response(cursor, "record")
        return ret if "by_id" in cond.keys() and cond["by_id"] else sorted(ret, key=lambda x: x["time_stamp"])

      
def query_display(q: dict[str, Any]) -> list[str]:
    """
    Query the display name for a given key.
    :param q: A dictionary containing the queries.
        Possible key:
            query (list[list[str, str]]): The queries.
                Each item in the list should be a list with exactly 2 elements, where the first element is the key
                and the second is the table in which the query should be made.
                Example:
                    {
                      "cond": {"query": [["example_tag", "tag"], ["science207", "classroom"]]}
                    }
    :return: A list of display names. Invalid queries return N/A.
    """
    with sqlite3.connect("database.db") as db:
        special_query_keys = ["tag", "place"]
        cursor = db.cursor()
        ret = []
        for query in q["query"]:
            key_name = "id"
            if query[1] in special_query_keys:
                key_name = query[1]
            cursor.execute("select * from " + query[1] + f" where {key_name}=:key",
                           {"key": query[0]})
            res = cursor.fetchall()
            if len(res) == 0:
                ret.append("N/A")
            for i in range(len(res)):
                item = res[i]
                query = q["query"][i]
                heads = cursor.execute(f"pragma table_info({query[1]})").fetchall()
                idx = -1
                for head in heads:
                    if head[1] == "display":
                        idx = head[0]
                        break
                ret.append(item[idx] if idx != -1 else "N/A")
        return ret


def check_permission(permissions: list[str], classrooms: list[str]) -> list[str]:
    """
    Check if the given classrooms can be reserved with presented permissions.
    :param permissions: Permission IDs
    :param classrooms: Classroom IDs
    :return: A list containing the classrooms that CANNOT be reserved with the given permissions.
    """
    with sqlite3.connect("database.db") as db:
        allowed_classrooms = []
        cursor = db.cursor()
        for permission in permissions:
            cursor.execute("select * from permission where id=:id", {"id": permission})
            res = cursor.fetchall()
            if len(res) == 0:
                continue
            allowed_classrooms.extend(res[0][2].split(",")[:-1])
        allowed_classrooms = list(set(allowed_classrooms))
        if "*" in allowed_classrooms:
            return []

        no_permissions = []
        for classroom in classrooms:
            if classroom not in allowed_classrooms:
                no_permissions.append(classroom)

        return no_permissions


def query_img(url: str) -> bytes:
    """
    Get the image with the specified url.
    TODO: Format should not be restricted to jpg, but it works for now.
    :param url: the url of the image
    :return: the image
    """
    with open(f"img/{url}.jpg", "rb") as img_file:
        img_data = img_file.read()
    return img_data


def query_user(cond: dict[str, Any]) -> list[dict[str, Any]]:
    """
    Get the info of the user.
    :param cond: A dictionary containing the filters that select the users to return.
        Possible keys:
            id (int): The user id.
            display (str): The user's display name.
            permission (list[str]): The user's permissions.
    :return: The users that satisfy the conditions of the filter.
    """
    with sqlite3.connect("database.db") as db:
        cursor = db.cursor()
        param = utils.construct_params(cond)
        if len(cond) != 0:
            cursor.execute("select * from user_info where " + utils.construct_condition(cond), param)
        else:
            cursor.execute("select * from user_info")
        ret = utils.construct_response(cursor, "user_info")
        return ret

      
def judge_conflict(classroom: str, noon: bool, time_stamp: int) -> bool:
    with sqlite3.connect("database.db") as db:
        cursor = db.cursor()
        cursor.execute("""
            SELECT COUNT(*) FROM record   
            WHERE classroom_id = :classroom AND noon = :noon AND time_stamp = :time_stamp
        """, {"classroom": classroom, "noon": noon, "time_stamp": time_stamp})
        count = cursor.fetchone()[0]
        return count > 0


def get_all_func_tags() -> dict[str, str]:
    """
    Get all available function tags and their display names.
    :return: Function tags (keys) and their display names (values).
    """
    with open("func_tags") as tag_file:
        tags = tag_file.read().split("\n")
        del tags[-1]
    q_list = [[tags[i], "tag"] for i in range(0, len(tags))]
    r = query_display({"query": q_list})

    return {tags[i]: r[i] for i in range(0, len(q_list))}


def get_cyclical(cond: dict[str, Any]) -> list[dict[str, str]]:
    with sqlite3.connect("database.db") as db:
        cursor = db.cursor()
        param = utils.construct_params(cond)
        cursor.execute("select * from cyclical_record where " + utils.construct_condition(cond), param)
        ret = utils.construct_response(cursor, "cyclical_record")
        return ret

def generate_schedule():
    """
    Get the schedule of current week.
    :return: A dictionary with the following keys:
                success (bool): Whether the reservation was successful.
                err_code (int)[optional]: The error code if an error occurs. Possible codes:
                    100: Python Exception
                error (str)[optional]: Error message.
    """
    try:
        on_duty = list(map(lambda x: x["display"],query_user({"permission": ["duty"]})))
        random.shuffle(on_duty)
        raw_records = query_record({})

        now = datetime.datetime.now()
        start = now - datetime.timedelta(days=now.weekday())
        start.replace(hour=0, minute=0, second=0, microsecond=0)
        end = start + datetime.timedelta(days=7)

        records: list[list[dict[str, Any]]] = [[] for _ in range(7)]
        for i in range(len(raw_records)):
            if start.timestamp() <= raw_records[i]["time_stamp"] <= end.timestamp():
                records[datetime.datetime.fromtimestamp(raw_records[i]["time_stamp"]).weekday()].append(raw_records[i])

        wb = Workbook()
        ws = wb.active

        normal_style = NamedStyle(name="normal")
        normal_style.font = Font(name="Times New Roman", bold=False, color="000000", size=8)
        normal_style.alignment = Alignment(horizontal="center", vertical="center")

        title_style = NamedStyle(name="title")
        title_style.font = Font(name="Times New Roman", bold=True, color="000000", size=16)
        title_style.alignment = Alignment(horizontal="center", vertical="center")

        ws["A1"] = (f"重庆南开中学学生社团中午活动安排表"
                    f"（{start.year}/{start.month}/{start.day}-{end.year}/{end.month}/{end.day}）")
        ws.merge_cells("A1:J5")
        title_cell = ws["A1"]
        title_cell.style = title_style
        weekdays = ["一", "二", "三", "四", "五"]
        for i in range(5):
            ws[f"{chr(i * 2 + ord('A'))}6"] = "星期" + weekdays[i]
            ws.merge_cells(f"{chr(i * 2 + ord('A'))}6:{chr(i * 2 + ord('B'))}7")
            ws[f"{chr(i * 2 + ord('A'))}6"].style = normal_style
            ws[f"{chr(i * 2 + ord('A'))}6"].font = Font(name="Times New Roman", bold=True, color="000000", size=11)

        max_item = max(len(records[i]) for i in range(len(records)))
        duty_cnt = 0
        for i in range(5):
            for j in range(max_item):
                if j < len(records[i]):
                    club_display = query_user({"id": records[i][j]["applicant_id"]})[0]["display"]
                    classroom_display = query_classroom({"id": records[i][j]["classroom_id"]})[0]["display"]
                    ws[f"{chr(i * 2 + ord('A'))}{8 + j * 4}"]\
                        = f"活动社团：{club_display}\n协助人：{on_duty[duty_cnt % len(on_duty)]}\n地点：{classroom_display}"
                    duty_cnt += 1
                ws.merge_cells(f"{chr(i * 2 + ord('A'))}{8 + j * 4}:{chr(i * 2 + ord('B'))}{11 + j * 4}")
                ws[f"{chr(i * 2 + ord('A'))}{8 + j * 4}"].style = normal_style

        wb.save("schedule.xlsx")
        wb.close()
        return {"success": True}
    except BaseException as e:
        traceback.print_exc()
        return {"success": False, "err_code": 100, "error": e}


def generate_statistics():
    with sqlite3.connect("database.db") as db:
        cursor = db.cursor()
        registered_users = cursor.execute("select * from user_info where register_num != 0").fetchall()
        displays, register_nums = [], []
        for registered_user in registered_users:
            displays.append(registered_user[1])
            register_nums.append(registered_user[3])
        plt.rcParams["font.family"] = "SimHei"
        plt.ylabel("历史预约次数")
        plt.bar(np.array(displays), np.array(register_nums), color="#82007E")
        plt.gca().yaxis.set_major_locator(plt.MaxNLocator(integer=True))
        plt.title("各社团活动教室预约情况")
        plt.savefig("statistics.png")


def get_permission_passwords() -> list[dict[str, Union[str, bool]]]:
    with open("password_perm_mapping") as ppm:
        return [{"password": mapping.split(' ')[0],
                "permission": mapping.split(' ')[1],
                "isDisposable": True if mapping.split(' ')[2] == "1" else False} for mapping in ppm.read().strip().splitlines()]
